#!蟒蛇3
# updateProduce.py - 更正农产品销售电子表格中的成本。

import openpyxl

wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb["Sheet"]

# 农产品种类及其最新价格
PRICE_UPDATES = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}

# TODO：遍历行并更新价格。
# 遍历各行并更新价格。
for row_num in range(2, sheet.max_row):
    product_name = sheet.cell(row=row_num, column=1).value

    if product_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[product_name]

print("Done...")

wb.save("updatedProduceSales.xlsx")
