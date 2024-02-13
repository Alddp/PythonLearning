import openpyxl
import os
from pathlib import Path
from openpyxl.utils import get_column_letter, column_index_from_string

path = Path(r"D:\Destok\Work\自律会文件")
file_name = "22网技1、22网技2、23软件1、23电信1、23电信2.xlsx"
# os.getcwd()
# os.chdir(path)
wb = openpyxl.load_workbook(path / file_name)
type(wb)
wb.sheetnames
sheet = wb["Sheet1"]

a = sheet["B1"]
b = sheet[1]
for i in b:
    print(i.value)

c = sheet['A']
get_column_letter(a.column)
column_index_from_string()
a.value  # 学期'
a.row  # 1
a.column  # 2
a.coordinate  # 'B1"

list(sheet.columns)[0]

b = sheet["A1":"D5"]  #
for row_of_cell_objects in b:
    for cell_obj in row_of_cell_objects:
        print(cell_obj.coordinate, cell_obj.value)
    print("--- END OF ROW ---")
sheet.max_row
sheet.max_column
