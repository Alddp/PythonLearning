# 1．从Excel电子表格中读取数据。
# 3．计算每个县的总人口。
# 4．输出结果。
# 这意味着代码需要执行以下操作。
import openpyxl
from pprint import pformat
from pathlib import Path
from openpyxl.utils import get_column_letter, column_index_from_string

# 1．用openpyxl 模块打开Excel文档并读取单元格。
# 2．计算所有普查区和人口数据，并将它保存到一个数据结构中。
# 3．利用pprint 模块，将该数据结构写入一个扩展名为.py的文本文件。
file_name = Path(Path.cwd() / "./censuspopdata.xlsx")
wb = openpyxl.load_workbook(file_name)
sheet = wb[wb.sheetnames[0]]


count_data = {}
print("Reading rows...")
# 获取第一列的value
# a1 = sheet["A1" : get_column_letter(sheet.max_column) + "1"]
# for row in a1:
#     for cell in row:
#         print(cell.value)


for row in range(2, sheet.max_row):
    state = sheet["B" + str(row)].value
    county = sheet["C" + str(row)].value
    pop = sheet["D" + str(row)].value

    # 确保此状态的key存在。
    count_data.setdefault(state, {})
    # 确保此州中此县的key存在。
    count_data[state].setdefault(county, {"tracts": 0, "pop": 0})
    # 每行代表一个人口普查区域，因此请递增 1。
    count_data[state][county]["tracts"] += 1
    # 在此人口普查区域中逐个增加县人口。
    count_data[state][county]["pop"] += int(pop)


# TODO：打开一个新的文本文件，并将count_data的内容写入其中。
print("Writing results...")
with open("census2010.json", "w") as result:
    result.write(pformat(count_data))
print("Done...")
